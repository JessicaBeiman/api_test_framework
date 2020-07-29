# encoding = 'utf-8'
# author = 'jessica'
import openpyxl
from openpyxl.styles import Border, Side, Font
import time


class ParseExcel(object):
    def __init__(self):
        self.workbook = None
        self.excel_file = None
        self.font = Font(color=None)
        self.color_dict = {'red': 'FFFF3030', 'green': 'FF008B00'}

    def load_workbook(self, excel_path):
        try:
            self.workbook = openpyxl.load_workbook(excel_path)
        except Exception as e:
            raise e
        self.excel_file = excel_path
        return self.workbook

    def get_sheet_by_name(self, sheet_name):
        try:
            sheet = self.workbook[sheet_name]
            return sheet
        except Exception as e:
            raise e

    def get_sheet_by_index(self, sheet_index):
        try:
            sheet_name = self.workbook.sheetnames[sheet_index]
        except Exception as e:
            raise e
        sheet = self.workbook[sheet_name]
        return sheet

    def get_total_rows_number(self, sheet):
        return sheet.max_row

    def get_total_cols_number(self, sheet):
        return sheet.max_column

    def get_start_row_number(self, sheet):
        return sheet.min_row

    def get_start_col_number(self, sheet):
        return sheet.min_column

    def get_single_row(self, sheet, row_no):
        try:
            return list(sheet.rows)[row_no-1]
        except Exception as e:
            raise e

    def get_single_column(self, sheet, col_no):
        try:
            return list(sheet.columns)[col_no-1]
        except Exception as e:
            raise e

    def get_value_in_cell(self, sheet, coordinate=None, row_no=None, col_no=None):
        try:
            if coordinate:
                return sheet[coordinate].value
            elif coordinate is None and row_no is not None and col_no is not None:
                return sheet.cell(row=row_no, column=col_no).value
            else:
                raise Exception('Argument exception!')
        except Exception as e:
            raise e
        # if coordinate is not None:
        #     try:
        #         return sheet[coordinate].value
        #     except Exception as e:
        #         raise e
        # elif coordinate is None and row_no is not None and col_no is not None:
        #     try:
        #         return sheet.cell(row=row_no, column=col_no).value
        #     except Exception as e:
        #         raise e
        # else:
        #     raise Exception('Argument exception!')

    def write_cell(self, sheet, content, coordinate=None, row_no=None, col_no=None, color=None):
        try:
            if coordinate:
                sheet[coordinate].value = content
                if color:
                    sheet[coordinate].font = color
                self.workbook.save(self.excel_file)
            elif coordinate is None and row_no is not None and col_no is not None:
                sheet.cell(row=row_no, column=col_no).value = content
                if color:
                    sheet.cell(row=row_no, column=col_no).font = color
                self.workbook.save(self.excel_file)
            else:
                raise Exception('Argument exception!')
        except Exception as e:
            raise e

    def write_current_time_in_cell(self, sheet, coordinate=None, row_no=None, col_no=None):
        current_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
        self.write_cell(sheet=sheet, content=current_time, coordinate=coordinate, row_no=row_no, col_no=col_no)


if __name__ == '__main__':
    pe = ParseExcel()
    pe.load_workbook(r'C:\PycharmProjects\api_test_framework\test_data\inter_test_data.xlsx')
    pe.get_sheet_by_name('API')
    print('用index获取sheet的名字：', pe.get_sheet_by_index(2))
    sheet_obj = pe.get_sheet_by_index(5)
    print(type(sheet_obj))
    print(pe.get_total_rows_number(sheet_obj))
    print(pe.get_total_cols_number(sheet_obj))
    print(pe.get_start_row_number(sheet_obj))
    print(pe.get_start_col_number(sheet_obj))
    print(pe.get_single_row(sheet_obj, 1))
    print(pe.get_single_column(sheet_obj, 4))
    print(pe.get_value_in_cell(sheet_obj, 'A1'))
    print(pe.get_value_in_cell(sheet_obj, row_no=2, col_no=5))
    pe.write_cell(sheet_obj, 'test', 'A2')
    pe.write_cell(sheet_obj, 'test2', row_no=3, col_no=4)
    pe.write_current_time_in_cell(sheet_obj, 'C2')
    pe.write_current_time_in_cell(sheet_obj, row_no=2, col_no=6)


